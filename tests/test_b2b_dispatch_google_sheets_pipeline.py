from __future__ import annotations

from datetime import date
import json
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook

from inventory_cover.b2b_dispatch_schemas import B2B_MASTER_HEADERS
from inventory_cover.cli import _build_source_pipeline_tasks, build_parser
from inventory_cover.config import B2BDispatchPipelineConfig
from inventory_cover.io.b2b_dispatch_google_sheets_io import read_b2b_dispatch_google_sheet
from inventory_cover.pipelines.b2b_dispatch_pipeline import B2BDispatchPipeline


AS_OF_DATE = date(2026, 6, 30)
SPREADSHEET_ID = "test-spreadsheet"
RK_SHEET = "RK PO 007GK"
CLICKTECK_SHEET = "CLICKTECK DISPATCH "
CLICKTECH_ALIAS_SHEET = "CLICKTECH DISPATCH"
ETRADE_SHEET = "ETRADE DISPATCH "

RK_HEADERS = [
    "APPOINTMENT ID",
    "INVOICE NO",
    "BOXES",
    "PO",
    "LOC.",
    "ASIN",
    "PO+ASIN",
    "SKU",
    "PO DATE",
    "PO QTY",
    "Dispatch Qty",
    "UNIT VALUE",
    "TOTAL VALUE",
    "DATE",
    "LOCATION",
]

CLICK_HEADERS = [
    "Appointment ID",
    "INVOICE NO",
    "BOXES",
    "PO",
    "",
    "ASIN",
    "PO+ASIN",
    "MODEL NAME",
    "PO DATE",
    "PO QTY",
    "Dispatch Qty",
    "UNIT PRICE",
    "PO DISPATCH VALUE",
    "Date",
    "Location",
]

ETRADE_HEADERS = [
    "APPOINTMENT ID",
    "INVOICE NO",
    "BOXES",
    "PO",
    "Ship to location",
    "ASIN",
    "PO+ASIN",
    "MODEL NAME",
    "PO DATE",
    "PO QTY",
    "Dispatch Qty",
    "UNIT VALUE",
    "TOTAL VALUE",
    "Date",
    "Location",
]


class FakeSheetsClient:
    def __init__(self, values_by_title: dict[str, list[list[Any]]]):
        self.values_by_title = values_by_title
        self.requested_ranges: list[str] = []

    def list_sheet_titles(self, spreadsheet_id: str) -> list[str]:
        assert spreadsheet_id == SPREADSHEET_ID
        return [*self.values_by_title, "BLINKIT DISPATCH", "INSTAMART DISPATCH"]

    def get_values(self, spreadsheet_id: str, range_name: str) -> list[list[Any]]:
        assert spreadsheet_id == SPREADSHEET_ID
        self.requested_ranges.append(range_name)
        return self.values_by_title[_title_from_range(range_name)]


def test_google_sheets_mode_produces_master_contract_and_inclusive_window(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    fake_client = FakeSheetsClient(
        {
            RK_SHEET: _sheet_values(
                3,
                RK_HEADERS,
                [
                    _row(po="PO-RK-START", dispatch_date="28-06-2026"),
                    _row(po="PO-RK-MID", dispatch_date="29-06-2026"),
                    _row(po="PO-RK-END", dispatch_date="30-06-2026"),
                    _row(po="PO-RK-OLD", dispatch_date="27-06-2026"),
                ],
            ),
            CLICKTECK_SHEET: _sheet_values(3, CLICK_HEADERS, [_row(po="PO-CLICK", dispatch_date="30-06-2026")]),
            ETRADE_SHEET: _sheet_values(4, ETRADE_HEADERS, [_row(po="PO-ETRADE", dispatch_date="28-06-2026")]),
        }
    )
    _patch_google_reader(monkeypatch, fake_client)

    result = B2BDispatchPipeline(_google_config(tmp_path)).run()
    records = _master_records(result.backend_output_file)
    metadata = json.loads(result.metadata_file.read_text(encoding="utf-8"))

    assert _master_headers(result.backend_output_file) == list(B2B_MASTER_HEADERS)
    assert {record["PO"] for record in records} == {
        "PO-RK-START",
        "PO-RK-MID",
        "PO-RK-END",
        "PO-CLICK",
        "PO-ETRADE",
    }
    assert {record["Source Sheet"] for record in records} == {RK_SHEET, CLICKTECK_SHEET, ETRADE_SHEET}
    assert "OUTSIDE_LOOKBACK_WINDOW" in _issue_types(result.backend_output_file)
    assert {_title_from_range(value) for value in fake_client.requested_ranges} == {
        RK_SHEET,
        CLICKTECK_SHEET,
        ETRADE_SHEET,
    }
    assert metadata["b2b_source_mode"] == "google_sheets"
    assert metadata["google_target_sheet_names_requested"] == [RK_SHEET, CLICKTECK_SHEET, ETRADE_SHEET]
    assert metadata["google_target_sheet_names_found"] == [RK_SHEET, CLICKTECK_SHEET, ETRADE_SHEET]
    assert metadata["google_header_rows_used"][RK_SHEET] == 3
    assert metadata["google_header_rows_used"][CLICKTECK_SHEET] == 3
    assert metadata["google_header_rows_used"][ETRADE_SHEET] == 4
    assert metadata["date_window_used"]["lookback_start"] == "2026-06-28"
    assert metadata["date_window_used"]["lookback_end"] == "2026-06-30"
    assert metadata["row_counts"]["included"] == 5
    assert metadata["row_counts"]["excluded"] == 1
    assert metadata["row_counts"]["rejected"] == 0


def test_google_sheets_mode_rejects_missing_critical_fields(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    fake_client = FakeSheetsClient(
        {
            RK_SHEET: _sheet_values(
                3,
                RK_HEADERS,
                [
                    _row(po="", dispatch_date="30-06-2026"),
                    _row(po="PO-VALID", dispatch_date="30-06-2026"),
                ],
            ),
            CLICKTECK_SHEET: _sheet_values(3, CLICK_HEADERS, []),
            ETRADE_SHEET: _sheet_values(4, ETRADE_HEADERS, []),
        }
    )
    _patch_google_reader(monkeypatch, fake_client)

    result = B2BDispatchPipeline(_google_config(tmp_path)).run()

    assert {record["PO"] for record in _master_records(result.backend_output_file)} == {"PO-VALID"}
    assert "MISSING_PO" in _issue_types(result.backend_output_file)
    metadata = json.loads(result.metadata_file.read_text(encoding="utf-8"))
    assert metadata["row_counts"]["rejected"] == 1


def test_google_sheet_alias_tolerates_clicktech_spelling(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    fake_client = FakeSheetsClient(
        {
            RK_SHEET: _sheet_values(3, RK_HEADERS, []),
            CLICKTECH_ALIAS_SHEET: _sheet_values(3, CLICK_HEADERS, [_row(po="PO-CLICKTECH")]),
            ETRADE_SHEET: _sheet_values(4, ETRADE_HEADERS, []),
        }
    )
    _patch_google_reader(monkeypatch, fake_client)

    result = B2BDispatchPipeline(_google_config(tmp_path)).run()
    records = _master_records(result.backend_output_file)

    assert records[0]["PO"] == "PO-CLICKTECH"
    assert records[0]["Source Sheet"] == CLICKTECH_ALIAS_SHEET
    assert records[0]["Source Channel"] == "CLICKTECH"


def test_run_full_inventory_cover_b2b_source_is_passed_to_pipeline(monkeypatch: Any) -> None:
    captured: dict[str, Any] = {}

    class FakeB2BDispatchPipeline:
        def __init__(self, config: B2BDispatchPipelineConfig):
            captured["source_mode"] = config.source_mode

        def run(self) -> Any:
            return SimpleNamespace(
                run_id="RUN1",
                rows_written=0,
                validation_issue_count=0,
                duplicate_count=0,
                backend_output_file=Path("backend.xlsx"),
                metadata_file=Path("metadata.json"),
                log_file=Path("log.txt"),
            )

    monkeypatch.setattr("inventory_cover.cli.B2BDispatchPipeline", FakeB2BDispatchPipeline)
    args = build_parser().parse_args(
        [
            "run-full-inventory-cover",
            "--skip-po-items",
            "--skip-sales-inventory",
            "--b2b-source",
            "google-sheets",
        ]
    )

    tasks = _build_source_pipeline_tasks(args)
    assert [task.command for task in tasks] == ["run-b2b-dispatch"]
    tasks[0].runner()
    assert captured["source_mode"] == "google_sheets"


def _patch_google_reader(monkeypatch: Any, fake_client: FakeSheetsClient) -> None:
    def fake_reader(config: B2BDispatchPipelineConfig, run_id: str) -> Any:
        return read_b2b_dispatch_google_sheet(config, run_id, client=fake_client)

    monkeypatch.setattr("inventory_cover.io.b2b_dispatch_source_provider.read_b2b_dispatch_google_sheet", fake_reader)


def _google_config(tmp_path: Path) -> B2BDispatchPipelineConfig:
    return B2BDispatchPipelineConfig(
        project_root=tmp_path,
        input_dir=tmp_path / "incoming",
        run_root=tmp_path / "runs",
        processed_dir=tmp_path / "processed" / "b2b_dispatch",
        as_of_date=AS_OF_DATE,
        lookback_days=2,
        source_mode="google_sheets",
        google_spreadsheet_id=SPREADSHEET_ID,
    )


def _sheet_values(header_row: int, headers: list[str], rows: list[dict[str, Any]]) -> list[list[Any]]:
    values: list[list[Any]] = [["summary/formula row"] for _ in range(header_row - 1)]
    values.append(headers)
    for row in rows:
        values.append([_value_for_source_header(header, row) for header in headers])
    return values


def _row(
    po: str = "PO-1",
    asin: str = "ASIN1",
    invoice_no: str = "INV1",
    dispatch_date: object = "30-06-2026",
    dispatch_qty: object = 5,
    unit_value: object = 100,
    dispatch_value_source: object = 500,
) -> dict[str, object]:
    return {
        "Appointment ID": "APT1",
        "Invoice No": invoice_no,
        "Boxes": 2,
        "PO": po,
        "Ship To Location": "BLR1",
        "ASIN": asin,
        "PO ASIN Key": f"{po}{asin}",
        "Model Number": "MODEL1",
        "PO Date": "29-06-2026",
        "PO Qty": 10,
        "Dispatch Qty": dispatch_qty,
        "Unit Value": unit_value,
        "Dispatch Value Source": dispatch_value_source,
        "Dispatch Date": dispatch_date,
        "Dispatch Location": "WH1",
    }


def _value_for_source_header(header: str, row: dict[str, Any]) -> object:
    source_to_canonical = {
        "APPOINTMENT ID": "Appointment ID",
        "Appointment ID": "Appointment ID",
        "INVOICE NO": "Invoice No",
        "BOXES": "Boxes",
        "PO": "PO",
        "LOC.": "Ship To Location",
        "Ship to location": "Ship To Location",
        "ASIN": "ASIN",
        "PO+ASIN": "PO ASIN Key",
        "SKU": "Model Number",
        "MODEL NAME": "Model Number",
        "PO DATE": "PO Date",
        "PO QTY": "PO Qty",
        "Dispatch Qty": "Dispatch Qty",
        "UNIT VALUE": "Unit Value",
        "UNIT PRICE": "Unit Value",
        "TOTAL VALUE": "Dispatch Value Source",
        "PO DISPATCH VALUE": "Dispatch Value Source",
        "DATE": "Dispatch Date",
        "Date": "Dispatch Date",
        "LOCATION": "Dispatch Location",
        "Location": "Dispatch Location",
    }
    if header == "":
        return None
    return row.get(source_to_canonical[header])


def _title_from_range(range_name: str) -> str:
    prefix = range_name.split("!", 1)[0]
    return prefix[1:-1].replace("''", "'")


def _master_headers(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    headers = [cell.value for cell in wb["B2B_Dispatch_Master"][1]]
    wb.close()
    return headers


def _master_records(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["B2B_Dispatch_Master"]
    headers = [cell.value for cell in ws[1]]
    records = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return records


def _issue_types(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Validation_Issues"]
    headers = [cell.value for cell in ws[1]]
    issue_index = headers.index("Issue Type")
    values = [row[issue_index] for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return values
