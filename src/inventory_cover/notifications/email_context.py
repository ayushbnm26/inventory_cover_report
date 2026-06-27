"""Build email-safe report context from Pipeline 4 run artifacts."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from inventory_cover.inventory_cover_schemas import (
    INVENTORY_SHEET,
    SALES_SHEET,
    InventoryCoverPipelineRunResult,
)


NOT_AVAILABLE = "Not available"

REPORT_CONTEXT_KEYS: tuple[str, ...] = (
    "sales_period_start",
    "sales_period_end",
    "sales_report_updated_date",
    "inventory_period_start",
    "inventory_period_end",
    "inventory_report_updated_date",
    "b2b_dispatch_as_of_date",
    "b2b_dispatch_lookback_start",
    "b2b_dispatch_lookback_end",
)


@dataclass(frozen=True)
class EmailReportContext:
    """Business context rendered into the report delivery email and audit JSON."""

    run_id: str
    generated_at: str
    product_count: int
    validation_issue_count: int
    warning_count: int
    team_workbook_path: Path
    backend_workbook_path: Path
    report_context: dict[str, str]
    missing_fields: tuple[str, ...]

    @property
    def team_workbook_name(self) -> str:
        return self.team_workbook_path.name


def build_email_report_context(
    result: InventoryCoverPipelineRunResult,
    *,
    logger: logging.Logger | None = None,
) -> EmailReportContext:
    """Extract stable report context from run-specific Pipeline 4 artifacts."""

    metadata = _read_metadata(result.metadata_file, logger=logger)
    generated_at = _format_value(metadata.get("end_time") or metadata.get("start_time"))
    if generated_at == NOT_AVAILABLE:
        generated_at = _format_value(_mtime(result.team_output_file))

    source_summaries = _read_source_summary_from_workbook(result.backend_output_file, logger=logger)
    if not source_summaries:
        source_summaries = _source_summaries_from_metadata(metadata)

    summaries = {str(row.get("Source Type") or row.get("source_type") or ""): row for row in source_summaries}
    sales = summaries.get("Sales", {})
    inventory = summaries.get("Inventory", {})
    b2b = summaries.get("B2B Dispatch", {})

    b2b_context = _read_b2b_run_summary(b2b, logger=logger)
    report_context = {
        "sales_period_start": _field(sales, "Report Period Start", "report_period_start"),
        "sales_period_end": _field(sales, "Report Period End", "report_period_end"),
        "sales_report_updated_date": _field(sales, "Report Updated Date", "report_updated_date"),
        "inventory_period_start": _field(inventory, "Report Period Start", "report_period_start"),
        "inventory_period_end": _field(inventory, "Report Period End", "report_period_end"),
        "inventory_report_updated_date": _field(inventory, "Report Updated Date", "report_updated_date"),
        "b2b_dispatch_as_of_date": _format_value(b2b_context.get("As of date")),
        "b2b_dispatch_lookback_start": _format_value(b2b_context.get("Lookback start date")),
        "b2b_dispatch_lookback_end": _format_value(b2b_context.get("Lookback end date")),
    }
    _fill_source_date_gaps(
        report_context,
        sales,
        sheet_name=SALES_SHEET,
        mapping={
            "sales_period_start": ("Viewing Range Start", "min"),
            "sales_period_end": ("Viewing Range End", "max"),
            "sales_report_updated_date": ("Report Updated Date", "max"),
        },
        logger=logger,
    )
    _fill_source_date_gaps(
        report_context,
        inventory,
        sheet_name=INVENTORY_SHEET,
        mapping={
            "inventory_period_start": ("Viewing Range Start", "min"),
            "inventory_period_end": ("Viewing Range End", "max"),
            "inventory_report_updated_date": ("Report Updated Date", "max"),
        },
        logger=logger,
    )
    missing_fields = tuple(key for key in REPORT_CONTEXT_KEYS if report_context[key] == NOT_AVAILABLE)
    if logger is not None:
        for field in missing_fields:
            logger.warning("Report context field unavailable: %s", field)

    return EmailReportContext(
        run_id=result.run_id,
        generated_at=generated_at,
        product_count=result.product_count,
        validation_issue_count=result.validation_issue_count,
        warning_count=result.warning_count,
        team_workbook_path=result.team_output_file,
        backend_workbook_path=result.backend_output_file,
        report_context=report_context,
        missing_fields=missing_fields,
    )


def _read_metadata(path: Path, *, logger: logging.Logger | None) -> dict[str, Any]:
    if not path.exists():
        _log_warning(logger, "Run metadata file not found: %s", path)
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        _log_warning(logger, "Could not read run metadata %s: %s", path, exc)
        return {}


def _read_source_summary_from_workbook(path: Path, *, logger: logging.Logger | None) -> list[dict[str, Any]]:
    if not path.exists():
        _log_warning(logger, "Backend audit workbook not found for email context: %s", path)
        return []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            if "Source_Summary" not in wb.sheetnames:
                _log_warning(logger, "Source_Summary sheet not found in backend audit workbook: %s", path)
                return []
            ws = wb["Source_Summary"]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
    except Exception as exc:  # noqa: BLE001 - context fallback should continue.
        _log_warning(logger, "Could not read Source_Summary from %s: %s", path, exc)
        return []

    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        records.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    return records


def _source_summaries_from_metadata(metadata: dict[str, Any]) -> list[dict[str, Any]]:
    summaries = metadata.get("source_summaries")
    if not isinstance(summaries, list):
        return []
    return [row for row in summaries if isinstance(row, dict)]


def _read_b2b_run_summary(summary: dict[str, Any], *, logger: logging.Logger | None) -> dict[str, Any]:
    for key in ("Copied Run Path", "copied_run_path", "Source Latest Path", "source_latest_path"):
        value = summary.get(key)
        if not value:
            continue
        path = Path(str(value))
        data = _read_b2b_run_summary_sheet(path, logger=logger)
        if data:
            return data
    return {}


def _fill_source_date_gaps(
    report_context: dict[str, str],
    summary: dict[str, Any],
    *,
    sheet_name: str,
    mapping: dict[str, tuple[str, str]],
    logger: logging.Logger | None,
) -> None:
    missing = {key: column_mode for key, column_mode in mapping.items() if report_context[key] == NOT_AVAILABLE}
    if not missing:
        return
    for key in ("Copied Run Path", "copied_run_path", "Source Latest Path", "source_latest_path"):
        value = summary.get(key)
        if not value:
            continue
        extracted = _read_source_date_fields(Path(str(value)), sheet_name, missing, logger=logger)
        for context_key, extracted_value in extracted.items():
            if report_context[context_key] == NOT_AVAILABLE:
                report_context[context_key] = extracted_value
        if all(report_context[context_key] != NOT_AVAILABLE for context_key in missing):
            return


def _read_source_date_fields(
    path: Path,
    sheet_name: str,
    mapping: dict[str, tuple[str, str]],
    *,
    logger: logging.Logger | None,
) -> dict[str, str]:
    if not path.exists():
        _log_warning(logger, "Source backend workbook not found for email context: %s", path)
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                _log_warning(logger, "%s sheet not found in source backend workbook: %s", sheet_name, path)
                return {}
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
    except Exception as exc:  # noqa: BLE001 - unavailable fallback context is non-blocking.
        _log_warning(logger, "Could not read %s from %s: %s", sheet_name, path, exc)
        return {}

    if len(rows) < 2:
        return {}
    headers = [str(value) if value is not None else "" for value in rows[0]]
    positions = {header: index for index, header in enumerate(headers)}
    extracted: dict[str, str] = {}
    for context_key, (column, mode) in mapping.items():
        index = positions.get(column)
        if index is None:
            continue
        values = [row[index] for row in rows[1:] if index < len(row) and _format_value(row[index]) != NOT_AVAILABLE]
        if values:
            extracted[context_key] = _format_value(_pick_value(values, mode))
    return extracted


def _read_b2b_run_summary_sheet(path: Path, *, logger: logging.Logger | None) -> dict[str, Any]:
    if not path.exists():
        _log_warning(logger, "B2B backend workbook not found for email context: %s", path)
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            if "Run_Summary" not in wb.sheetnames:
                _log_warning(logger, "Run_Summary sheet not found in B2B backend workbook: %s", path)
                return {}
            ws = wb["Run_Summary"]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
    except Exception as exc:  # noqa: BLE001 - unavailable B2B context is non-blocking.
        _log_warning(logger, "Could not read B2B Run_Summary from %s: %s", path, exc)
        return {}

    if len(rows) < 2:
        return {}
    headers = [str(value) if value is not None else "" for value in rows[0]]
    values = rows[1]
    return {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}


def _field(row: dict[str, Any], workbook_key: str, metadata_key: str) -> str:
    return _format_value(row.get(workbook_key, row.get(metadata_key)))


def _format_value(value: Any) -> str:
    if value is None:
        return NOT_AVAILABLE
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else NOT_AVAILABLE


def _pick_value(values: list[Any], mode: str) -> Any:
    dated = [value for value in values if isinstance(value, (datetime, date))]
    if dated:
        return min(dated) if mode == "min" else max(dated)
    return values[0] if mode == "min" else values[-1]


def _mtime(path: Path) -> datetime | None:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime)
    except OSError:
        return None


def _log_warning(logger: logging.Logger | None, message: str, *args: Any) -> None:
    if logger is not None:
        logger.warning(message, *args)
